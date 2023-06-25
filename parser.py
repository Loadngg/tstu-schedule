from re import search
from openpyxl import load_workbook
from utils import *
from typing import List
from customtkinter import CTkLabel


class Parser:
    def __init__(self, info_label: CTkLabel) -> None:
        self.books: List[openpyxl.Workbook] = []
        self.info_label = info_label

    def load_files(self, files: tuple) -> None:
        self.info_label.configure(text="Загрузка...")

        for item in files:
            print(item)
            book = open_xls_as_xlsx(item) \
                if item[-5:].split('.')[1] == 'xls' \
                else load_workbook(item)

            self.books.append(book)

        self.info_label.configure(text="Загрузка завершена")

    def search(self, search_filter: str) -> List[str]:
        self.info_label.configure(text="Обработка...")

        result: List[str] = []

        for book in self.books:
            for sheet_name in book.get_sheet_names():
                sheet = book.get_sheet_by_name(sheet_name)

                time_column_exist = search("время", str(sheet.cell(row=1, column=2).value))

                for cellObj in sheet[1:sheet.max_row]:
                    for cell in cellObj:
                        if cell.value is not None and str(cell.value) != '' and search(search_filter, str(cell.value)):
                            row_index = cell.row

                            while sheet.cell(row=row_index, column=1).value is None:
                                row_index -= 1

                            result.append(
                                ' '.join(
                                    [
                                        remove_extra_whitespaces(sheet.cell(row=row_index, column=1).value)[:5],
                                        ' '.join(
                                            [
                                                remove_extra_whitespaces(sheet.cell(row=cell.row, column=2).value),
                                                remove_extra_whitespaces(
                                                    sheet.cell(row=cell.row + 1, column=2).value)
                                            ]
                                        ) if time_column_exist else '',
                                        remove_extra_whitespaces(sheet.cell(row=1, column=cell.column).value),
                                        remove_extra_whitespaces(cell.value),
                                    ]
                                )
                            )

            print(f"Book {self.books.index(book)} processed")

        self.info_label.configure(text="Обработка завершена")

        return result
