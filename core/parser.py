import openpyxl

from re import search
from typing import List

from PyQt5.QtWidgets import QLabel
from openpyxl import load_workbook
from fuzzywuzzy import fuzz
from fuzzywuzzy import process

from core.utils import remove_extra_whitespaces, open_xls_as_xlsx, info_output


class Parser:
    def __init__(self, info_label: QLabel = None) -> None:
        self.books: List[openpyxl.Workbook] = []
        self.books_path: List[str] = []
        self.info_label = info_label

    def clear_books(self) -> None:
        self.books.clear()
        self.books_path.clear()

    def load_files(self, files: List[str]) -> List[openpyxl.Workbook]:
        info_output(self.info_label, "Загрузка...")

        for item in files:
            if item in self.books_path:
                continue
            book = open_xls_as_xlsx(item) \
                if item[-5:].split('.')[1] == 'xls' \
                else load_workbook(item)

            self.books_path.append(item)
            self.books.append(book)
            print(f"Loaded {item}")

        info_output(self.info_label, f"Успешно загружено файлов: {self.books.__len__()}")

        return self.books

    def search(self, search_filter: str) -> List[str]:
        info_output(self.info_label, "Обработка...")

        result: List[str] = []

        for book in self.books:
            for sheet_name in book.get_sheet_names():
                sheet = book.get_sheet_by_name(sheet_name)

                time_column_exist = search("время", str(sheet.cell(row=1, column=2).value))

                for cellObj in sheet[1:sheet.max_row]:
                    for cell in cellObj:
                        if (cell.value is not None and str(cell.value) != '' and
                                fuzz.WRatio(search_filter.lower(), str(cell.value).lower()) >= 75):
                            row_index = cell.row

                            while sheet.cell(row=row_index, column=1).value is None:
                                row_index -= 1

                            time_with_pare = True if sheet.cell(row=cell.row + 1, column=2).value is None else False

                            result.append(
                                ' '.join(
                                    [
                                        remove_extra_whitespaces(sheet.cell(row=row_index, column=1).value)[:5],
                                        remove_extra_whitespaces(
                                            ' '.join(
                                                [
                                                    remove_extra_whitespaces(
                                                        sheet.cell(row=cell.row, column=2).value).replace(',', ''),
                                                    '' if time_with_pare else remove_extra_whitespaces(
                                                        sheet.cell(row=cell.row + 1, column=2).value)
                                                ]
                                            ) if time_column_exist else ''
                                        ),
                                        remove_extra_whitespaces(sheet.cell(row=1, column=cell.column).value),
                                        remove_extra_whitespaces(cell.value),
                                    ]
                                )
                            )

            print(f"Book {self.books.index(book)} processed")

        info_output(self.info_label, "Обработка завершена")

        return result
